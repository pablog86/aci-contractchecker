#!/usr/bin/env python3
# To create Virtual ENV
# python3 -m venv contractchecker_env
# source contractchecker_env/bin/activate
# **********************************************************************************
# **********************************************************************************
# Script desarrollado por Pablo Gomez - CCIE DC 57661
# **********************************************************************************
# **********************************************************************************

import json
import requests
from requests.packages.urllib3.exceptions import InsecureRequestWarning
import warnings
import collections.abc
import sys
import os
import re
import numpy as np
from itertools import cycle
import inspect
from collections import Counter
from datetime import datetime
import envs

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

warnings.simplefilter('ignore', InsecureRequestWarning)

# **********************************************************************************
# MODIFIABLE DEFINITIONS
# **********************************************************************************

verify_https = False  # Validate https certificate
page_size = 2000  # Elements per page (low numbers may generate some issues)

# ----------------------------------------------------------------------

priorities = {"class-eq-filter": 1,
              "class-eq-deny": 2,
              "class-eq-allow": 3,
              "prov-nonshared-to-cons": 4,
              "black_list": 5,
              "fabric_infra": 6,
              "fully_qual": 7,
              "system_incomplete": 8,
              "src_dst_any": 9,
              "shsrc_any_filt_perm": 10,
              "shsrc_any_any_perm": 11,
              "shsrc_any_any_deny": 12,
              "src_any_filter": 13,
              "any_dest_filter": 14,
              "src_any_any": 15,
              "any_dest_any": 16,
              "any_any_filter": 17,
              "grp_src_any_any_deny": 18,
              "grp_any_dest_any_deny": 19,
              "grp_any_any_any_permit": 20,
              "any_any_any": 21,
              "any_vrf_any_deny": 22
              }

# **********************************************************************************
# System Reserved pcTag - This pcTag is used for system internal rules (1-15).
# Globally scoped pcTag - This pcTag is used for shared service (16-16385).
# Locally scoped pcTag - This pcTag is locally used per VRF (range from 16386-65535).
# **********************************************************************************

pctags = {
    "0": "any",
    "1": "ignore",  # Ignore in Spine-Proxy, ARP, Multicast, etc.
    "10": "pctag10-unknown",
    "13": "ext-shrsvc",
    "14": "int-shrsvc",
    "15": "ext-0.0.0.0/0"  # External EPG 0.0.0.0/0
}

# ----------------------------------------------------------------------

_debug = 0
_debugLog = False
_color_i = "\033[0;33;40m"
_color_f = "\033[0m"


def debug(obj, msj="debug msj", level=3):
    if _debug >= level:
        print(
            _color_i +
            str(msj) +
            "\n{}".format(
                json.dumps(
                    obj,
                    indent=4)) +
            _color_f)
        if _debugLog:
            with open(os.path.join(sys.path[0], "debuglog.json"), "a") as debugfile:
                debugfile.write(str(msj) + "\n")
                json.dump(obj, indent=4, sort_keys=True, fp=debugfile)
                debugfile.write("\n")

# --------


def printt(string=None):
    if string:
        print(str(string))
    if _debugLog:
        with open(os.path.join(sys.path[0], "debuglog.json"), "a") as debugfile:
            debugfile.write(str(string) + "\n")

# --------


def update(d, u) -> dict:
    for k, v in u.items():
        if isinstance(v, collections.abc.Mapping):
            d[k] = update(d.get(k, {}), v)
        else:
            d[k] = v
    return d


# --------


def count_elem(l, maxcount) -> int:
    # Sum the count of repeated elements
    s = sum(Counter([list(elem)[0] for elem in l]).values())
    try:
        # The number of the most concurrent element
        m = max(Counter([list(elem)[0] for elem in l]).values())
    except ValueError:
        m = 0
        debug("count_elem() - > empty max()", level=2)
    aux = m if m == maxcount else s
    debug(
        aux,
        "{} response aggregated lenght:".format(
            inspect.stack()[1][3]),
        2)
    return aux

# ---------------------------------------------------------------------------------------------------------------------------------------------
# API Calls
# ---------------------------------------------------------------------------------------------------------------------------------------------

# Login to APIC #
w = cycle(("|", "/", "-", "\\"))
envs.TOKEN = None

def apic_login() -> str:
    print('Working: (%s)\r' % next(w), end="")
    token = ""
    err = ""
    try:
        response = requests.post(
            url=envs.APIC_URL + "/api/aaaLogin.json",
            headers={
                "Content-obj": "application/json; charset=utf-8",
            },
            data=json.dumps(
                {
                    "aaaUser": {
                        "attributes": {
                            "name": envs.USERNAME,
                            "pwd": envs.PASS
                        }
                    }
                }
            ),
            verify=False)
        json_response = json.loads(response.content)
        token = json_response['imdata'][0]['aaaLogin']['attributes']['token']
    except KeyError:
        print(
            "HTTP Request failed, Status Code: {status_code}".format(
                status_code=response.status_code))
    except:
        print("Connection error.")
        sys.exit()
    return token


# ---------------------------------------------------------------------------------------------------------------------------------------------


def get_method(
        url,
        query_target=None,
        target_subtree_class=None,
        query_target_filter=None,
        page=0) -> list:
    #token = apic_login()
    try:
        response = requests.get(
            url,
            headers={
                "Cookie": "APIC-cookie=" + envs.TOKEN,
                "Content-obj": "application/json; charset=utf-8",
            },
            params={
                "query-target": query_target,
                "target-subtree-class": target_subtree_class,
                "query-target-filter": query_target_filter,
                "page-size": page_size,
                "page": page},
            verify=verify_https)
        debug(
            response.status_code,
            "Debug output CODE {} -> (url={}, query_target={}, target-subtree-class={}, query-target-filter={}, page-size={}, page={}): ".format(
                inspect.stack()[1][3],
                url,
                query_target,
                target_subtree_class,
                query_target_filter,
                page_size,
                page),
            1)
        if response.status_code == requests.codes.ok:
            debug(
                response.json(),
                "Debug output {} -> (url={}, query_target={}, target-subtree-class={}, query-target-filter={}, page-size={}, page={}): ".format(
                    inspect.stack()[1][3],
                    url,
                    query_target,
                    target_subtree_class,
                    query_target_filter,
                    page_size,
                    page))
            debug(response.json()["totalCount"], "totalCount: ", 1)
            return response
        else:
            return None
    except requests.exceptions.RequestException:
        debug(
            "HTTP Request failed, Status Code: {status_code}".format(
                status_code=response.status_code))
        return None


# ---------------------------------------------------------------------------------------------------------------------------------------------
# Output formatting
# ---------------------------------------------------------------------------------------------------------------------------------------------


def printable(d_contract, wr=False):
    if not bool(d_contract):
        printt("No matching criteria -> empty output")
        return None
# ------------------------------------
    try:
        printt(
            "\n################### Contract: {} ###################\n".format(
                d_contract["dn"]))
        printt("Consumers: ")
        for i in d_contract["Consumers"]:
            printt(i)
        printt("----------------------------------------")
        printt("Providers: ")
        for i in d_contract["Providers"]:
            printt(i)
        printt("----------------------------------------")
        printt("Subjects: ")
        for k, v in d_contract["Subjects"].items():
            printt("{}:{}".format(k, v))
        printt("----------------------------------------")
        printt("Contract inheritance is omitted")
        printt("----------------------------------------")
    except KeyError:
        pass
# ------------------------------------
    contract_list = pd.DataFrame(columns=['id', 'Source', 'Destination', 'VRF', 'Contract ', 'Filter', 'Action', 'Prio', 'Priority', 'Direction', 'State'])
    i = 1
    for rules in d_contract:
        for rule in d_contract[rules]:
            r = d_contract[rules][rule]
            contract_list.loc[i] = [r["id"], 
            r["sPcTag-str"].replace("uni/tn-", "").replace("/ap-", "/").replace("/epg-", "/").replace("/out-", "/").replace("/instP-", "/").replace("/ldevCtx-c-","/").replace("/ctx-", "/"), 
            r["dPcTag-str"].replace("uni/tn-", "").replace("/ap-", "/").replace("/epg-", "/").replace("/out-", "/").replace("/instP-", "/").replace("/ldevCtx-c-","/").replace("/ctx-", "/"), 
            r["scopeId-str"].replace("uni/tn-", "").replace("/ctx-", "/"), 
            r["fltName"].replace("uni/tn-", "").replace("/bcr-","/"), 
            r["fltId"], 
            r["action"], 
            priorities[r["prio"]], 
            r["prio"], 
            r["direction"], 
            r["operSt"]]
            i += 1
    contract_list = contract_list.set_index("id")
    contract_list = contract_list.sort_values(by="Prio", ascending=True)
    printt(contract_list.to_markdown(tablefmt="psql")) #tablefmt="grid"
    if wr:
        wb = openpyxl.Workbook()
        wbname = str(next(iter(d_contract))).replace("/","_")+".xlsx"
        printt("Writing {} on same directory".format(wbname))
        with pd.ExcelWriter(wbname, engine="openpyxl") as writer:
            contract_list.to_excel(writer, sheet_name="ACL_Rules", startrow=0)
    printt(datetime.now())
# ---------------------------------------------------------------------------------------------------------------------------------------------
# VRFs class
# ---------------------------------------------------------------------------------------------------------------------------------------------


class VRFs (object):
    __objTypeVrf = "fvCtx"

    def __init__(self, filters=None):
        self. filters = filters
        self.d_vrfs = {}

        self.get_vrf()


    def get_vrf(self):
        vrfs = self.get_node_objs(self.__objTypeVrf, self.filters)  # fvACtx
        if vrfs is None:
            return None
        for vrf in vrfs:
            if "fvCtx" in vrf:
                obj_id = "fvCtx"
                ctx_id = "dn"
                scope_id = "scope"
                pctag = "pcTag"
            elif "fvCtxDef" in vrf:
                obj_id = "fvCtxDef"
                ctx_id = "ctxDn"
                scope_id = "scope"
                pctag = "pcTag"
            elif "fvTnlCtx" in vrf:
                obj_id = "fvTnlCtx"
                ctx_id = "dn"
                scope_id = "scope"
                pctag = "pcTag"
            else:
                continue
            self.d_vrfs.update(
                {vrf[obj_id]["attributes"][scope_id]: vrf[obj_id]["attributes"][ctx_id]})
            self.d_vrfs.update(
                {vrf[obj_id]["attributes"][ctx_id]: vrf[obj_id]["attributes"][scope_id]})
            self.d_vrfs.update(
                {"{}-pctag".format(vrf[obj_id]["attributes"][ctx_id]): vrf[obj_id]["attributes"][pctag]})
            if int(vrf[obj_id]["attributes"][pctag]) < 16386:
                self.d_vrfs.update(
                    {vrf[obj_id]["attributes"][pctag]: "{}-pctag".format(vrf[obj_id]["attributes"][ctx_id])})
        self.d_vrfs.update({"16777200": "uni/tn-infra/black-hole"})
        #self.d_vrfs.update({"16777199": "uni/tn-infra/ap-access/epg-default"})
        #self.d_vrfs.update({"1": "uni/tn-mgmt/extmgmt-default"})
        debug(self.d_vrfs, "VRFs: ", 2)


    def get_node_objs(self, obj, filters=None) -> list:
        print('Working: (%s)\r' % next(w), end="")
        url = envs.APIC_URL + "/api/node/class/{}.json".format(obj)
        response = get_method(url, query_target_filter=filters)
        if response is not None:
            aux = response.json()["imdata"]
            i = 1
            while count_elem(
                aux, int(
                    response.json()["totalCount"])) < int(
                    response.json()["totalCount"]):  # len(aux)
                response = get_method(url, query_target_filter=filters, page=i)
                aux = aux + response.json()["imdata"]
                i = i + 1
            debug(len(aux), "get_node_objs response lenght:", 1)
            if count_elem(
                aux, int(
                    response.json()["totalCount"])) > int(
                    response.json()["totalCount"]):
                printt(
                    "More elements ({}) than totalCount ({})".format(
                        aux, response.json()["totalCount"]))
            return aux
        else:
            return []

# ---------------------------------------------------------------------------------------------------------------------------------------------
# EPGs class
# ---------------------------------------------------------------------------------------------------------------------------------------------


class EPGs (VRFs):

    __epgsFilterType = {
        "ctx": "ctxDefDn",
        "epg": "epgPKey",
        "scope": "scopeId",
        "bd": "dn"}

    __objTypeEpgs = ("fvAREpP", "vzToEPg", "fvBD", "vnsEPgDef")

    def __init__(self, filters=[]):
        VRFs.__init__(self)
        self.filters = filters
        self.d_epgs = {}
        self.epgs()


    def epgs(self):
        if len(self.filters) > 0:
            for filte in self.filters:
                f = filte.split("/")
                f = f[-1].split("-")[0]
                if f == "scope":
                    filt = filte[6:]
                for epg_t in self.__objTypeEpgs:
                    self.mapping_epg_pctag(
                        epg_t, "wcard({}.{}, \"{}\")".format(
                            epg_t, self.__epgsFilterType[f], filt))
        else:
            for epg_t in self.__objTypeEpgs:
                self.mapping_epg_pctag(epg_t)


    def mapping_epg_pctag(self, obj, filters=None):

        epgs = self.get_node_objs(obj, filters)
        if obj == "fvAREpP":
            l3outsAny = self.get_l3extsubnet(
                "eq(l3extSubnet.ip, \"0.0.0.0/0\")")
        if epgs is None:
            return

        if not bool(self.d_vrfs):  # Check empty dict
            return
        for epg in epgs:
            if "fvEpP" in epg:
                obj_id = "fvEpP"
                ctx_id = "scopeId"
                epg_id = "epgPKey"
            elif "vzToEPg" in epg:
                obj_id = "vzToEPg"
                ctx_id = "scopeId"
                epg_id = "epgDn"
            elif "fvRtdEpP" in epg:  # instP
                obj_id = "fvRtdEpP"
                ctx_id = "scopeId"
                epg_id = "epgPKey"
            elif "fvInBEpP" in epg:
                obj_id = "fvInBEpP"
                ctx_id = "scopeId"
                epg_id = "epgPKey"
            elif "fvOoBEpP" in epg:
                obj_id = "fvOoBEpP"
                ctx_id = "scopeId"
                epg_id = "epgPKey"
            elif "fvBD" in epg:
                obj_id = "fvBD"
                ctx_id = "scope"
                epg_id = "dn"
            elif "fvBDDef" in epg:
                obj_id = "fvBDDef"
                ctx_id = "scope"
                epg_id = "bdDn"
            elif "fvAEPg" in epg:
                obj_id = "fvAEPg"
                ctx_id = "scope"
                epg_id = "dn"
            elif "vnsEPgDef" in epg:
                obj_id = "vnsEPgDef"
                ctx_id = "dn"
                epg_id = "lIfCtxDn"
            else:
                continue

            if obj_id == "vnsEPgDef":  # Service Graph shadow EPG
                try:
                    if int(epg[obj_id]["attributes"]["pcTag"]) > 16385:  # pcTag local
                        ctx_name = re.findall(
                            r"(?<=S-\[).+?(?=\])",
                            epg[obj_id]["attributes"][ctx_id])[0]  # TODO: some SG doesnt inform ctx
                        self.d_epgs.update({self.d_vrfs[ctx_name]: ctx_name})
                        if ctx_name in self.d_epgs:
                            self.d_epgs[ctx_name].update(
                                {epg[obj_id]["attributes"]["pcTag"]: epg[obj_id]["attributes"][epg_id]})
                        else:
                            self.d_epgs[ctx_name] = {
                                epg[obj_id]["attributes"]["pcTag"]: epg[obj_id]["attributes"][epg_id]}
                            self.d_epgs[ctx_name].update(
                                {self.d_vrfs["{}-pctag".format(ctx_name)]: ctx_name})
                    else: # pcTag global
                        self.d_epgs.update(
                                {epg[obj_id]["attributes"]["pcTag"]: str(epg[obj_id]["attributes"][epg_id])[:-18]})
                except KeyError:
                    printt(
                        "Undef scope:{} -> sg object: {} ".format(
                            ctx_name,
                            epg[obj_id]["attributes"][epg_id]))

            else:
                try:
                    self.d_epgs.update(
                        {epg[obj_id]["attributes"][ctx_id]: self.d_vrfs[epg[obj_id]["attributes"][ctx_id]]})
                    if self.d_vrfs[epg[obj_id]["attributes"]
                                   [ctx_id]] in self.d_epgs:
                        self.d_epgs[self.d_vrfs[epg[obj_id]["attributes"][ctx_id]]].update(
                            {epg[obj_id]["attributes"]["pcTag"]: epg[obj_id]["attributes"][epg_id]})
                    else:
                        self.d_epgs[self.d_vrfs[epg[obj_id]["attributes"][ctx_id]]] = {
                            epg[obj_id]["attributes"]["pcTag"]: epg[obj_id]["attributes"][epg_id]}
                        self.d_epgs[self.d_vrfs[epg[obj_id]["attributes"][ctx_id]]].update({self.d_vrfs["{}-pctag".format(
                            self.d_vrfs[epg[obj_id]["attributes"][ctx_id]])]: self.d_vrfs[epg[obj_id]["attributes"][ctx_id]]})
                    if obj_id == "fvRtdEpP":
                        for l3outAny in l3outsAny:
                            if re.search(
                                    epg[obj_id]["attributes"][epg_id],
                                    l3outAny["l3extSubnet"]["attributes"]["dn"]):
                                self.d_epgs[self.d_vrfs[epg[obj_id]["attributes"][ctx_id]]].update(
                                    {"15": "{}(0.0.0.0/0)".format(epg[obj_id]["attributes"][epg_id])})
                except KeyError:
                    printt(
                        "Undef scope:{} -> object: {} ".format(
                            epg[obj_id]["attributes"][ctx_id],
                            epg[obj_id]["attributes"][epg_id]))
                try:
                    if int(epg[obj_id]["attributes"]["pcTag"]) < 16386:  # pcTag global
                        self.d_epgs.update(
                            {epg[obj_id]["attributes"]["pcTag"]: epg[obj_id]["attributes"][epg_id]})
                except ValueError:
                    pass

        self.d_epgs.update({"16777200": self.d_vrfs["16777200"]})  # black-hole


    def get_l3extsubnet(self, filters) -> list:
        url = envs.APIC_URL + \
            "/api/node/class/l3extSubnet.json"
        response = get_method(
            url,
            query_target_filter=filters)
        if response is not None:
            aux = response.json()["imdata"]
            i = 1
            while count_elem(
                aux, int(
                    response.json()["totalCount"])) < int(
                    response.json()["totalCount"]):  # len(aux)
                response = get_method(
                    url,
                    query_target_filter=filters,
                    page=i)
                aux = aux + response.json()["imdata"]
                i = i + 1
            debug(len(aux), "get_l3extsubnet response lenght:", 1)
            if count_elem(
                aux, int(
                    response.json()["totalCount"])) > int(
                    response.json()["totalCount"]):
                printt(
                    "More elements ({}) than totalCount ({})".format(
                        aux, response.json()["totalCount"]))
            return aux
        else:
            return []

# ---------------------------------------------------------------------------------------------------------------------------------------------
# Contracts class
# ---------------------------------------------------------------------------------------------------------------------------------------------


class Contracts (EPGs):

    __rtype = ("implicit", "implarp", "default")

    def __init__(self, pod_id, node_id, tenant=None, contract=None):
        self.pod_id = pod_id
        self.node_id = node_id
        self.tenant = tenant
        self.contract = contract

        self.urlfilterinfo = envs.APIC_URL + \
            "/api/node/class/topology/pod-{}/node-{}/vzRsRFltAtt.json".format(self.pod_id, self.node_id)
        self.urlzoningrule = envs.APIC_URL + \
            "/api/node/class/topology/pod-{}/node-{}/actrlRule.json".format(self.pod_id, self.node_id)
        self.urlcontract = envs.APIC_URL + \
            "/api/node/mo/uni/tn-{}/brc-{}.json".format(self.tenant, self.contract)
        self.urlsubject = envs.APIC_URL + \
            "/api/node/mo/uni/tn-{}/brc-{}/".format(self.tenant, self.contract)

        self.__brc = "uni/tn-{}/brc-{}".format(tenant, contract)
        self.filters = []
        self.d_contract = {}

        self.zoningrules = []
        self.__scopes = []

        self.mapping_zoningrule_contract()

        if bool(self.d_contract):
            self.node = "rules/pod-{}/node-{}".format(
                self.pod_id, self.node_id)
            for i in self.d_contract[self.node]:
                self.__scopes.append(self.d_contract[self.node][i]["scopeId"])
            self.__scopes = set(self.__scopes)

        if tenant is None or contract is None:
            EPGs.__init__(self)
        else:
            for scope in self.__scopes:
                self.filters.append("scope-{}".format(scope))
            EPGs.__init__(self, self.filters)
        self.contract_rules()
        debug(self.d_vrfs, "VRFs: ", 2)
        debug(self.d_epgs, "EPGs: ", 2)
        debug(self.d_contract, "Contracts: ", 2)


    def contract_rules(self):

        if not bool(self.d_contract):
            return

        if self.tenant is None or self.contract is None:
            d_fltInfo = self.get_contracts_info(self.urlfilterinfo)
        else:
            d_fltInfo = self.get_contracts_info(
                self.urlfilterinfo,
                filters="wcard(vzRsRFltAtt.dn, \"{}\")".format(
                    self.__brc))
        debug(d_fltInfo, "Filter Info: ", 3)
        for i in self.d_contract[self.node]:
            self.d_contract[self.node][i]["scopeId-str"] = self.d_epgs[self.d_contract[self.node][i]["scopeId"]]

            if self.d_contract[self.node][i]["sPcTag"] != "any":
                if self.d_contract[self.node][i]["sPcTag"] in pctags:  # reserved pcTag
                    try:
                        self.d_contract[self.node][i]["sPcTag-str"] = self.d_epgs[self.d_contract[self.node][i]["scopeId-str"]]["15"]
                    except KeyError:
                        self.d_contract[self.node][i]["sPcTag-str"] = pctags[self.d_contract[self.node][i]["sPcTag"]]
                else:
                    try:
                        if self.d_contract[self.node][i]["sPcTag"].isdigit() and int(
                                self.d_contract[self.node][i]["sPcTag"]) < 16386:  # pcTag Global
                            self.d_contract[self.node][i]["sPcTag-str"] = self.d_epgs[self.d_contract[self.node][i]["sPcTag"]]
                        else:
                            self.d_contract[self.node][i]["sPcTag-str"] = self.d_epgs[self.d_contract[self.node]
                                                                                  [i]["scopeId-str"]][self.d_contract[self.node][i]["sPcTag"]]
                    except KeyError:
                        self.d_contract[self.node][i]["sPcTag-str"] = self.d_vrfs[self.d_contract[self.node][i]["sPcTag"]]

            if self.d_contract[self.node][i]["dPcTag"] != "any":
                if self.d_contract[self.node][i]["dPcTag"] in pctags:   # reserved pcTag
                    try:
                        self.d_contract[self.node][i]["dPcTag-str"] = self.d_epgs[self.d_contract[self.node][i]["scopeId-str"]]["15"]
                    except KeyError:
                        self.d_contract[self.node][i]["dPcTag-str"] = pctags[self.d_contract[self.node][i]["dPcTag"]]
                else:
                    try:
                        if self.d_contract[self.node][i]["dPcTag"].isdigit() and int(
                                self.d_contract[self.node][i]["dPcTag"]) < 16386:  # pcTag Global
                            self.d_contract[self.node][i]["dPcTag-str"] = self.d_epgs[self.d_contract[self.node][i]["dPcTag"]]
                        else:
                            self.d_contract[self.node][i]["dPcTag-str"] = self.d_epgs[self.d_contract[self.node]
                                                                                  [i]["scopeId-str"]][self.d_contract[self.node][i]["dPcTag"]]
                    except KeyError:
                        self.d_contract[self.node][i]["dPcTag-str"] = self.d_vrfs[self.d_contract[self.node][i]["dPcTag"]]

            ##Add info to Contract Name ----  
            for fltInfo in d_fltInfo:
                f = fltInfo["vzRsRFltAtt"]["attributes"]["dn"]
                
                sTag = self.d_contract[self.node][i]["sPcTag-str"].replace("-pctag", "")
                if sTag in self.d_epgs.keys():#== self.d_contract[self.node][i]["scopeId-str"]:
                    sTag = self.d_epgs[sTag]["15"]
                elif sTag == pctags["15"]:
                    sTag = self.d_epgs[self.d_contract[self.node][i]["scopeId-str"]]["15"]
                
                dTag = self.d_contract[self.node][i]["dPcTag-str"].replace("-pctag", "")
                if dTag in self.d_epgs.keys():#== self.d_contract[self.node][i]["scopeId-str"]: 
                    dTag = self.d_epgs[dTag]["15"]
                elif dTag == pctags["15"]:
                    try:
                        dTag = self.d_epgs[self.d_contract[self.node][i]["scopeId-str"]]["15"]  
                    except:
                        dTag = "instP-" #Leaking mix the 15 pctag in a different VRF, TODO: improve this
                
                try:
                    if re.search(sTag.replace("(0.0.0.0/0)", ""), f) and re.search(dTag.replace("(0.0.0.0/0)", ""), f) and re.search(self.d_contract[self.node][i]["fltName"], f):
                        aux = re.findall(
                            r"(?<=/cdef-\[).+?(?=\])",
                            fltInfo["vzRsRFltAtt"]["attributes"]["dn"])[0]
                        if "GraphInst_C-[" in aux:
                            aux = re.findall(r"(?<=/GraphInst_C-\[).+?(?=\])",aux)[0]
                        if aux is not None:
                            self.d_contract[self.node][i]["fltName"] = aux
                except BaseException:
                    debug(
                        "re.search error (bad character range n-a), {} sPcTag: {},  dPcTag: {}, fltId: {}".format(
                            self.d_contract[self.node][i]["id"],
                            self.d_contract[self.node][i]["sPcTag-str"],
                            self.d_contract[self.node][i]["dPcTag-str"],
                            self.d_contract[self.node][i]["fltId"]),
                        level=1)


        # Default filter management purge | TODO: improve this
        if self.tenant is not None or self.contract is not None:
            r = []
            for i in self.d_contract[self.node]:
                if self.d_contract[self.node][i]["fltName"] != self.__brc:
                    r.append(i)
            for i in r:
                del self.d_contract[self.node][i]


    def mapping_zoningrule_contract(self):

        rule_type = (
            "id",
            "sPcTag",
            "dPcTag",
            "fltId",
            "direction",
            "operSt",
            "scopeId",
            "action",
            "prio",
            )
        rules = {}
# All filters in the switch
        if self.tenant is None or self.contract is None:  
            self.zoningrules = self.get_contracts_info(self.urlzoningrule)
            for zoningrule in self.zoningrules:
                rule = zoningrule["actrlRule"]["attributes"]["dn"]
                rules.update({rule: {}})
                for t in rule_type:
                    rules[rule].update(
                        {t: zoningrule["actrlRule"]["attributes"][t]})

                rules[rule].update({"scopeId-str":"any"})
                rules[rule].update({"sPcTag-str":"any"})
                rules[rule].update({"dPcTag-str":"any"})

                if zoningrule["actrlRule"]["attributes"]["ctrctName"] != "":
                    try:
                        if ":" in zoningrule["actrlRule"]["attributes"]["ctrctName"]:
                            aux = zoningrule["actrlRule"]["attributes"]["ctrctName"].split(":")
                            rules[rule].update({"fltName": "uni/tn-{}/brc-{}".format(aux[0], aux[1])})
                        else:
                            rules[rule].update({"fltName": zoningrule["actrlRule"]["attributes"]["ctrctName"]})
                    except IndexError:
                        rules[rule].update({"fltName": zoningrule["actrlRule"]["attributes"]["fltId"]})  #To Default filter management
                else:
                    rules[rule].update({"fltName": zoningrule["actrlRule"]["attributes"]["fltId"]}) #To Default filter management

            self.d_contract.update({"rules/pod-{}/node-{}".format(self.pod_id, self.node_id): rules})
# Filters matching the tenant/contract
        else:  
            self.get_contract()
            self.zoningrules = self.get_contracts_info(
                self.urlzoningrule, filters="wcard(actrlRule.ctrctName,\"{}:{}\")".format(
                    self.tenant, self.contract))
            if self.zoningrules == []:
                self.zoningrules = self.get_contracts_info(
                    self.urlzoningrule, filters="wcard(actrlRule.fltId,\"default\")")
            if self.zoningrules != []:
                for zoningrule in self.zoningrules:
                    rule = zoningrule["actrlRule"]["attributes"]["dn"]
                    rules.update({rule: {}})
                    for t in rule_type:
                        rules[rule].update(
                            {t: zoningrule["actrlRule"]["attributes"][t]})

                    if zoningrule["actrlRule"]["attributes"]["ctrctName"] != "":
                        try:
                            if ":" in zoningrule["actrlRule"]["attributes"]["ctrctName"]:
                                aux = zoningrule["actrlRule"]["attributes"]["ctrctName"].split(":")
                                rules[rule].update({"fltName": "uni/tn-{}/brc-{}".format(aux[0], aux[1])})
                            else:
                                rules[rule].update({"fltName": zoningrule["actrlRule"]["attributes"]["ctrctName"]})       
                        except IndexError:
                            rules[rule].update({"fltName": zoningrule["actrlRule"]["attributes"]["fltId"]}) #To Default filter management
                    else:
                        rules[rule].update({"fltName": zoningrule["actrlRule"]["attributes"]["fltId"]}) #To Default filter management
                    #rules[rule].update({"fltName": zoningrule["actrlRule"]["attributes"]["fltId"]}) #To Default filter management

                self.d_contract.update({"rules/pod-{}/node-{}".format(self.pod_id, self.node_id): rules})
            else:
                self.d_contract.update({"rules/pod-{}/node-{}".format(self.pod_id, self.node_id): {}})


    def get_contract(self):
        contracts = self.get_contracts_info(self.urlcontract)
        if bool(contracts):
            self.d_contract = {
                "dn": contracts[0]["vzBrCP"]["attributes"]["dn"]}
            self.d_contract.update({"Consumers": []})
            for c in self.get_contracts_info(
                    self.urlcontract, "children", "vzRtCons"):
                self.d_contract["Consumers"].append(
                    c["vzRtCons"]["attributes"]["tDn"])
            self.d_contract.update({"Providers": []})
            for c in self.get_contracts_info(
                    self.urlcontract, "children", "vzRtProv"):
                self.d_contract["Providers"].append(
                    c["vzRtProv"]["attributes"]["tDn"])
            self.d_contract.update({"Subjects": {}})
            for c in self.get_contracts_info(
                    self.urlcontract, "children", "vzSubj"):
                self.d_contract["Subjects"].update(
                    {c["vzSubj"]["attributes"]["dn"]: []})
                for s in self.get_contracts_info(
                        self.urlsubject,
                        query="children",
                        subtree="vzRsSubjFiltAtt",
                        subject=c["vzSubj"]["attributes"]["name"]):
                    self.d_contract["Subjects"][c["vzSubj"]["attributes"]["dn"]].append(
                        s["vzRsSubjFiltAtt"]["attributes"]["tDn"])

    def get_contracts_info(
            self,
            url,
            query=None,
            subtree=None,
            filters=None,
            subject=None) -> list:
        if subject:
            url = url + "subj-{}.json".format(subject)
        response = get_method(
            url,
            query_target=query,
            target_subtree_class=subtree,
            query_target_filter=filters)
        if response is not None:
            aux = response.json()["imdata"]
            i = 1
            while count_elem(
                aux, int(
                    response.json()["totalCount"])) < int(
                    response.json()["totalCount"]):  # len(aux)
                response = get_method(
                    url,
                    query_target=query,
                    target_subtree_class=subtree,
                    query_target_filter=filters,
                    page=i)
                aux = aux + response.json()["imdata"]
                i = i + 1
            debug(len(aux), "get_contracts_info response lenght:", 1)
            if count_elem(
                aux, int(
                    response.json()["totalCount"])) > int(
                    response.json()["totalCount"]):
                printt(
                    "More elements ({}) than totalCount ({})".format(
                        aux, response.json()["totalCount"]))
            return aux
        else:
            return []


# ---------------------------------------------------------------------------------------------------------------------------------------------

def getNode1Ver (token) -> bool:
    try:
        response = requests.get(
            url=envs.APIC_URL + "/api/node/class/topology/pod-1/node-1/firmwareCtrlrRunning.json",
            headers={
                "Cookie": "APIC-cookie=" + token,
                "Content-obj": "application/json; charset=utf-8",
            },
            verify=False)
        if response.status_code == 200:
            json_response = json.loads(response.content)
            version = json_response['imdata'][0]['firmwareCtrlrRunning']['attributes']['version']
            return version
    except KeyError:
        print(
            "HTTP Request failed, Status Code: {status_code}".format(
                status_code=response.status_code))
    
# ---------------------------------------------------------------------------------------------------------------------------------------------
# MAIN args parser
# ---------------------------------------------------------------------------------------------------------------------------------------------
# **********************************************************************************
# Create an envs.py file with (eg):
# URL="https://sandboxapicdc.cisco.com"
# envs.USERNAME = "admin"
# envs.PASS = "ciscopsdt"
# If there isn't a envs.py you can introduce the values in runtime
# **********************************************************************************

if __name__ == "__main__":
    import argparse
    from getpass import getpass
    no_envs = False
    try:
        import envs
    except ModuleNotFoundError:
        printt("No envs file located")
        no_envs = True

    separator = "-" * 110
    desc = separator + \
        """\nThis script generates a correlated output from the zoning-rule in the desired leaf switch, it runs locally
and using the APIC's APIs. It also has the posibility to filter the tenant/contract construct to validate
the correct renderization of the policy.
    """

    debugmsg = """Optional argument: debug level:
-d 1 = Response codes
-d 2 = Internal objs
-d 3 = Verbose"""

    def indent_formatter(prog): return argparse.RawTextHelpFormatter(
        prog, max_help_position=50)
    parser = argparse.ArgumentParser(
        prog="contract-checker",
        description=desc,
        formatter_class=indent_formatter,
        epilog=separator)

    parser.add_argument(
        'pod',
        metavar='podID',
        help='Pod ID number, eg: 1',
        type=int)
    parser.add_argument(
        'node',
        metavar='nodeID',
        help='Node ID number, eg: 101',
        type=int)
    parser.add_argument(
        '-t',
        '--tenant',
        action='store',
        help='Optional argument: Tenant of the contract to filter',
        metavar='Tenant Name')
    parser.add_argument(
        '-c',
        '--contract',
        action='store',
        help='Optional argument: contract to filter',
        metavar='Contract Name')
    parser.add_argument(
        '-d',
        '--debug',
        action='store',
        help=debugmsg,
        metavar='debug',
        type=int)
    parser.add_argument(
        '-l',
        '--logfile',
        action='store_true',
        help='Optional argument: log in a file')
    parser.add_argument(
        '-w',
        '--write',
        action='store_true',
        help='Optional argument: Write output to Excel')    

    args = parser.parse_args()

    if no_envs:
        envs.APIC_URL = str(input("APIC's URL: "))
        envs.APIC_URL = "https://{}".format(
            envs.APIC_URL) if envs.APIC_URL[0:4] != "http" else envs.APIC_URL
        envs.USERNAME = str(input("Username: "))
        envs.PASS = getpass("Password: ")
    else:
        try:
            envs.APIC_URL = envs.APIC_URL
        except AttributeError:
            envs.APIC_URL = str(input("APIC's URL: "))
            envs.APIC_URL = "https://{}".format(
                envs.APIC_URL) if envs.APIC_URL[0:4] != "http" else envs.APIC_URL
        try:
            envs.USERNAME = envs.USERNAME
        except AttributeError:
            envs.USERNAME = str(input("Username: "))
        try:
            envs.PASS = envs.PASS
        except AttributeError:
            envs.PASS = getpass("Password: ")

    _debug = args.debug if args.debug else _debug
    _debugLog = args.logfile
    print(envs.APIC_URL)
    print(envs.USERNAME)

    envs.TOKEN = apic_login()
    version = getNode1Ver(envs.TOKEN)
    print ("APIC1 version:",version)
    if version< "4.1":
        print("Unsupported APIC version")
        sys.exit()

    try:
        if args.tenant is None and args.contract is None:
            contract = Contracts(args.pod, args.node)
            printable(contract.d_contract, args.write)
        else:
            contract = Contracts(args.pod,
                                 args.node,
                                 args.tenant,
                                 args.contract)
            printable(contract.d_contract, args.write)

    except KeyboardInterrupt:
        printt("KeyboardInterrupt -> Goodbye!")
